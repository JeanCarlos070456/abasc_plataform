#!/usr/bin/env python
"""
Migração da base histórica de associados da ABASC.

Local esperado no projeto:
    membros/migracao.py
    membros/data.xlsx

Dependência:
    openpyxl

Fluxo recomendado:
    1) Apenas analisar a planilha (não grava nada):
       python membros/migracao.py

    2) Analisar apontando outra planilha:
       python membros/migracao.py --planilha "caminho/arquivo.xlsx"

    3) Gravar os registros válidos no banco Django/Supabase PostgreSQL:
       python membros/migracao.py --commit

    4) Preencher campos vazios de usuários que já existem:
       python membros/migracao.py --commit --update-existing

    5) Sobrescrever também campos já preenchidos (use com muita cautela):
       python membros/migracao.py --commit --update-existing --overwrite

IMPORTANTE
---------
- Este script NÃO cria senha temporária.
- Este script NÃO envia convite/e-mail em massa.
- Este script NÃO cria usuários no Supabase Auth automaticamente.
- Ele migra o cadastro para o User do Django, que no projeto ABASC usa o
  PostgreSQL configurado em DATABASE_URL (Supabase em produção).
- Para associados importados, supabase_user_id fica vazio. O fluxo de
  "Primeiro acesso" poderá criar/vincular o Supabase Auth somente quando
  o próprio associado solicitar o acesso.
- Todos os importados entram com role='associate'.
- O script é idempotente: procura usuário existente por e-mail, CPF e
  matrícula antes de criar outro registro.
- Por segurança, o modo padrão é somente análise. É obrigatório usar
  --commit para gravar no banco.
- Mesmo com --commit, apenas PRONTO_PARA_MIGRAR é gravado.
  PRECISA_REVISAO, SEM_EMAIL_VALIDO e DADO_CONFLITANTE são ignorados.
- Matrícula ambígua nunca é escolhida por adivinhação: ela fica vazia e
  registrada em "avisos" para revisão administrativa.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
    from openpyxl.utils.datetime import from_excel
except ImportError as exc:  # pragma: no cover - mensagem amigável no ambiente real
    raise SystemExit(
        "Dependência ausente: instale 'openpyxl' e adicione-a ao requirements.txt."
    ) from exc


# -----------------------------------------------------------------------------
# Caminhos / configuração
# -----------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
DEFAULT_XLSX = SCRIPT_DIR / "data.xlsx"
DEFAULT_REPORT = SCRIPT_DIR / "relatorio_migracao.csv"
DEFAULT_SETTINGS_MODULE = "abasc_mvp1.settings"

# A aba ADIMPLENTES é a base mais completa e contém o estado atual em sua
# coluna de situação. As demais são usadas para complementar e auditar.
SOURCE_PRIORITY = {
    "ADIMPLENTES": 100,
    "ASSOCIADOS ATIVOS": 90,
    "FALTA DE PAGAMENTO": 80,
    "INADIMPLENTES": 70,
    "VOLUTÁRIOS": 60,
    "NÃO CONCLUINTE": 50,
    "DIRETORIA": 40,
}

EXPECTED_SHEETS = tuple(SOURCE_PRIORITY)

STATUS_ACTIVE = "active"
STATUS_PENDING = "pending"
STATUS_OVERDUE = "overdue"
STATUS_INACTIVE = "inactive"

MIGRATION_READY = "PRONTO_PARA_MIGRAR"
MIGRATION_REVIEW = "PRECISA_REVISAO"
MIGRATION_CONFLICT = "DADO_CONFLITANTE"
MIGRATION_NO_EMAIL = "SEM_EMAIL_VALIDO"
MIGRATION_EXISTING = "JA_EXISTE_NO_SISTEMA"
MIGRATION_CREATED = "CRIADO"
MIGRATION_UPDATED = "ATUALIZADO"
MIGRATION_SKIPPED = "IGNORADO"
MIGRATION_ERROR = "ERRO"

EMPTY_MARKERS = {
    "",
    "-",
    "--",
    "N/A",
    "NA",
    "NÃO INFORMADO",
    "NAO INFORMADO",
    "NÃO SE APLICA",
    "NAO SE APLICA",
    "NONE",
    "NULL",
}

EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")

# Sanidade da validação de e-mail. Se esta regra for alterada incorretamente,
# o script falha antes de classificar toda a base como SEM_EMAIL_VALIDO.
if not EMAIL_RE.fullmatch("teste@example.com") or EMAIL_RE.fullmatch("n/a"):
    raise RuntimeError("Falha interna na validação de e-mail da migração.")


# Variações de nome confirmadas nesta base histórica.
KNOWN_NAME_ALIASES = {
    "ANDREA VILLA BOAS MELLO": "ANDREA VILLAS BOAS MELLO",
    "INAJARA CAGLIANI FERNANDES": "INAJARA CAGLIARI FERNANDES",
    "CAIO WILLIAN": "CAIO WILLIAM BATISTA DOS SANTOS",
    "ELIDA": "ELIDA DIAS CANDIDO",
    "LORENA FREITAS": "LORENA DE FREITAS FIDYK",
    "LEONARDO DE SOUZA LOURENCO": "LEONARDO DE SOUZA LOURENCO CARVALHO",
}

# Cabeçalhos aparecem com grafias diferentes entre as abas.
HEADER_ALIASES = {
    "membership_number": {
        "MATRICULA",
        "NUMERO DE ASSOCIADO",
        "NUMERO DO ASSOCIADO",
    },
    "full_name": {
        "NOME COMPLETO",
        "NOME",
    },
    "birth_date": {
        "DATA DE NASCIMENTO",
        "NASCIMENTO",
    },
    "cpf": {"CPF"},
    "gender": {"SEXO", "GENERO"},
    "city": {
        "MUNICIPIO DE RESIDENCIA",
        "MUNICIPIO",
        "CIDADE",
    },
    "state": {"UF", "ESTADO"},
    "phone": {
        "TELEFONE",
        "TELEFONE DE CONTATO",
        "WHATSAPP",
    },
    "email": {
        "E MAIL DE CONTATO",
        "EMAIL DE CONTATO",
        "E MAIL",
        "EMAIL",
    },
    "education_level": {
        "NIVEL DE ESCOLARIDADE",
        "ESCOLARIDADE",
    },
    "member_category": {"CATEGORIA"},
    "university": {
        "UNIVERSIDADE FACULDADE DE CONCLUSAO DA GRADUACAO",
        "UNIVERSIDADE FACULDADE",
        "UNIVERSIDADE",
        "FACULDADE",
    },
    "annuity_valid_until": {
        "VENCIMENTO ANUIDADE",
        "VALIDADE DA ANUIDADE",
        "VENCIMENTO DA ANUIDADE",
    },
}


# -----------------------------------------------------------------------------
# Estruturas de dados
# -----------------------------------------------------------------------------
@dataclass
class RawMember:
    source_sheet: str
    source_row: int
    source_priority: int
    status: str | None = None
    membership_number: str = ""
    full_name: str = ""
    birth_date: date | None = None
    cpf: str = ""
    cpf_raw: str = ""
    cpf_valid: bool = False
    gender: str = ""
    city: str = ""
    state: str = ""
    phone: str = ""
    email: str = ""
    email_valid: bool = False
    education_level: str = ""
    member_category: str = ""
    university: str = ""
    annuity_valid_until: date | None = None
    notes: list[str] = field(default_factory=list)

    @property
    def normalized_name(self) -> str:
        return canonical_person_name(self.full_name)


@dataclass
class ConsolidatedMember:
    rows: list[RawMember] = field(default_factory=list)
    membership_number: str = ""
    full_name: str = ""
    birth_date: date | None = None
    cpf: str = ""
    cpf_raw: str = ""
    cpf_valid: bool = False
    gender: str = ""
    city: str = ""
    state: str = ""
    phone: str = ""
    email: str = ""
    email_valid: bool = False
    education_level: str = ""
    member_category: str = ""
    university: str = ""
    annuity_valid_until: date | None = None
    association_status: str = STATUS_ACTIVE
    warnings: list[str] = field(default_factory=list)
    identity_conflict: bool = False
    migration_state: str = MIGRATION_READY
    db_action: str = ""
    db_user_id: str = ""

    @property
    def source_sheets(self) -> str:
        return ", ".join(sorted({row.source_sheet for row in self.rows}))


# -----------------------------------------------------------------------------
# Normalização / validação
# -----------------------------------------------------------------------------
def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = re.sub(r"\s+", " ", str(value)).strip()
    if normalize_key(text) in {normalize_key(v) for v in EMPTY_MARKERS}:
        return ""
    return text


def normalize_key(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper().strip()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def canonical_person_name(value: Any) -> str:
    """Retorna a chave canônica usada na consolidação da pessoa."""
    key = normalize_key(value)
    return KNOWN_NAME_ALIASES.get(key, key)


def normalize_email(value: Any) -> tuple[str, bool]:
    email = clean_text(value).lower().replace(" ", "")
    if not email:
        return "", False
    return email, bool(EMAIL_RE.match(email))


def only_digits(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\D", "", str(value))


def cpf_is_valid(value: Any) -> bool:
    cpf = only_digits(value)
    if len(cpf) != 11 or cpf == cpf[0] * 11:
        return False

    total = sum(int(cpf[i]) * (10 - i) for i in range(9))
    digit1 = (total * 10 % 11) % 10
    if digit1 != int(cpf[9]):
        return False

    total = sum(int(cpf[i]) * (11 - i) for i in range(10))
    digit2 = (total * 10 % 11) % 10
    return digit2 == int(cpf[10])


def normalize_cpf(value: Any) -> tuple[str, str, bool]:
    raw = clean_text(value)
    digits = only_digits(value)
    valid = cpf_is_valid(digits)
    if valid:
        formatted = f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"
        return formatted, raw, True
    return "", raw, False


def normalize_phone(value: Any) -> str:
    # Não inventamos DDI. O associado poderá revisar no onboarding.
    digits = only_digits(value)
    return digits[:20]


def normalize_state(value: Any) -> str:
    state = normalize_key(value).replace(" ", "")
    if state in {"NA", "N/A"}:
        return ""
    return state[:2] if len(state) >= 2 else ""


def normalize_membership_number(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    if isinstance(value, (int, float)):
        try:
            return str(int(value))
        except (ValueError, TypeError):
            pass
    digits = only_digits(text)
    return digits or text[:30]


def normalize_gender(value: Any) -> str:
    key = normalize_key(value)
    if key.startswith("FEM"):
        return "female"
    if key.startswith("MAS"):
        return "male"
    if "NAO BIN" in key or "NÃO BIN" in key:
        return "non_binary"
    return clean_text(value).lower()


def normalize_category(value: Any, source_sheet: str = "") -> str:
    key = normalize_key(value)
    if "VOLUNT" in key or source_sheet == "VOLUTÁRIOS":
        return "volunteer"
    if "JUNIOR" in key:
        return "junior"
    if "PLENO" in key:
        return "full"
    if not key:
        return ""
    return clean_text(value).lower()


def excel_date(value: Any, epoch: Any = None) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            dt = from_excel(value, epoch=epoch) if epoch is not None else from_excel(value)
            return dt.date() if isinstance(dt, datetime) else dt
        except Exception:
            return None

    text = clean_text(value)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def split_name(full_name: str) -> tuple[str, str]:
    parts = [part for part in clean_text(full_name).split(" ") if part]
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0].title(), ""
    return " ".join(parts[:-1]).title(), parts[-1].title()


def similar_names(name_a: str, name_b: str) -> bool:
    a = canonical_person_name(name_a)
    b = canonical_person_name(name_b)
    if not a or not b:
        return False
    return SequenceMatcher(None, a, b).ratio() >= 0.86


# -----------------------------------------------------------------------------
# Leitura da planilha
# -----------------------------------------------------------------------------
def canonical_header(value: Any) -> str | None:
    key = normalize_key(value)
    if not key:
        return None

    for canonical, aliases in HEADER_ALIASES.items():
        if key in aliases:
            return canonical

    # Alguns cabeçalhos carregam espaços/pontuação extras.
    if "NOME COMPLETO" in key:
        return "full_name"
    if "DATA DE NASCIMENTO" in key:
        return "birth_date"
    if key == "CPF" or key.startswith("CPF "):
        return "cpf"
    if key == "UF":
        return "state"
    if "MUNICIPIO" in key or key == "CIDADE":
        return "city"
    if "EMAIL" in key or "E MAIL" in key:
        return "email"
    if "TELEFONE" in key or "WHATSAPP" in key:
        return "phone"
    if "ESCOLARIDADE" in key:
        return "education_level"
    if key.startswith("CATEGORIA"):
        return "member_category"
    if "UNIVERSIDADE" in key or "FACULDADE" in key:
        return "university"
    if "VENCIMENTO" in key and "ANUIDADE" in key:
        return "annuity_valid_until"
    if "VALIDADE" in key and "ANUIDADE" in key:
        return "annuity_valid_until"
    if "MATRICULA" in key:
        return "membership_number"
    if key in {"SEXO", "GENERO"}:
        return "gender"
    return None


def find_header(ws) -> tuple[int, dict[str, int]]:
    best_row = 0
    best_map: dict[str, int] = {}

    for row_number in range(1, min(ws.max_row, 8) + 1):
        mapping: dict[str, int] = {}
        for col_number in range(1, min(ws.max_column, 60) + 1):
            canonical = canonical_header(ws.cell(row=row_number, column=col_number).value)
            if canonical and canonical not in mapping:
                mapping[canonical] = col_number

        # Nome é obrigatório para reconhecermos a linha de cabeçalho.
        if "full_name" in mapping and len(mapping) > len(best_map):
            best_row = row_number
            best_map = mapping

    if not best_row:
        raise ValueError(f"Não foi possível localizar o cabeçalho da aba '{ws.title}'.")
    return best_row, best_map


def infer_status(sheet_name: str, row_values: list[Any], name_col: int) -> str:
    sheet_name = sheet_name.upper()

    if sheet_name == "ADIMPLENTES":
        # Nesta aba, a coluna imediatamente anterior ao nome guarda
        # ADIMPLENTE/INADIMPLENTE e não possui cabeçalho confiável.
        candidate = ""
        if name_col > 1 and len(row_values) >= name_col - 1:
            candidate = normalize_key(row_values[name_col - 2])
        if "INADIMPLENTE" in candidate:
            return STATUS_OVERDUE
        if "ADIMPLENTE" in candidate:
            return STATUS_ACTIVE
        return STATUS_ACTIVE

    if sheet_name == "ASSOCIADOS ATIVOS":
        return STATUS_ACTIVE
    if sheet_name == "FALTA DE PAGAMENTO":
        return STATUS_PENDING
    if sheet_name == "INADIMPLENTES":
        return STATUS_OVERDUE
    if sheet_name == "NÃO CONCLUINTE":
        return STATUS_INACTIVE
    if sheet_name in {"VOLUTÁRIOS", "DIRETORIA"}:
        return STATUS_ACTIVE
    return STATUS_ACTIVE


def read_sheet(ws, epoch: Any) -> list[RawMember]:
    header_row, columns = find_header(ws)
    rows: list[RawMember] = []
    name_col = columns["full_name"]

    for row_number in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row=row_number, column=col).value for col in range(1, ws.max_column + 1)]

        def value_for(field: str) -> Any:
            col = columns.get(field)
            return ws.cell(row=row_number, column=col).value if col else None

        full_name = clean_text(value_for("full_name"))
        if not full_name:
            continue

        email, email_valid = normalize_email(value_for("email"))
        cpf, cpf_raw, cpf_valid = normalize_cpf(value_for("cpf"))
        birth = excel_date(value_for("birth_date"), epoch)
        annuity = excel_date(value_for("annuity_valid_until"), epoch)

        raw = RawMember(
            source_sheet=ws.title,
            source_row=row_number,
            source_priority=SOURCE_PRIORITY.get(ws.title, 0),
            status=infer_status(ws.title, values, name_col),
            membership_number=normalize_membership_number(value_for("membership_number")),
            full_name=full_name,
            birth_date=birth,
            cpf=cpf,
            cpf_raw=cpf_raw,
            cpf_valid=cpf_valid,
            gender=normalize_gender(value_for("gender")),
            city=clean_text(value_for("city")).title(),
            state=normalize_state(value_for("state")),
            phone=normalize_phone(value_for("phone")),
            email=email,
            email_valid=email_valid,
            education_level=clean_text(value_for("education_level")),
            member_category=normalize_category(value_for("member_category"), ws.title),
            university=clean_text(value_for("university")),
            annuity_valid_until=annuity,
        )

        if cpf_raw and not cpf_valid:
            raw.notes.append("CPF ausente ou inválido")
        if email and not email_valid:
            raw.notes.append("E-mail inválido")
        if not email:
            raw.notes.append("E-mail não informado")
        rows.append(raw)

    return rows


def load_members(path: Path) -> list[RawMember]:
    if not path.exists():
        raise FileNotFoundError(
            f"Planilha não encontrada: {path}\n"
            "Coloque o arquivo em membros/data.xlsx ou use --planilha."
        )

    wb = load_workbook(path, data_only=True, read_only=False)
    found = set(wb.sheetnames)
    missing = [name for name in EXPECTED_SHEETS if name not in found]
    if missing:
        print("[AVISO] Abas esperadas que não foram encontradas:", ", ".join(missing))

    all_rows: list[RawMember] = []
    for sheet_name in EXPECTED_SHEETS:
        if sheet_name not in found:
            continue
        sheet_rows = read_sheet(wb[sheet_name], wb.epoch)
        print(f"[LEITURA] {sheet_name}: {len(sheet_rows)} registro(s)")
        all_rows.extend(sheet_rows)
    return all_rows


# -----------------------------------------------------------------------------
# Consolidação / conflitos
# -----------------------------------------------------------------------------
def detect_cpf_collisions(rows: Iterable[RawMember]) -> set[str]:
    names_by_cpf: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        if row.cpf_valid and row.cpf:
            names_by_cpf[row.cpf].add(row.normalized_name)

    collisions = {
        cpf for cpf, names in names_by_cpf.items()
        if len({n for n in names if n}) > 1
    }
    return collisions


def detect_membership_collisions(rows: Iterable[RawMember]) -> set[str]:
    names_by_number: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        if row.membership_number:
            names_by_number[row.membership_number].add(row.normalized_name)
    return {
        number for number, names in names_by_number.items()
        if len({n for n in names if n}) > 1
    }


def choose_field(rows: list[RawMember], field_name: str) -> Any:
    # A matrícula na aba ASSOCIADOS ATIVOS deve ser preferida quando existir.
    if field_name == "membership_number":
        membership_rows = sorted(
            rows,
            key=lambda r: (
                1 if r.source_sheet == "ASSOCIADOS ATIVOS" else 0,
                r.source_priority,
            ),
            reverse=True,
        )
        for row in membership_rows:
            value = getattr(row, field_name)
            if value not in (None, ""):
                return value
        return ""

    ordered = sorted(rows, key=lambda r: r.source_priority, reverse=True)
    for row in ordered:
        value = getattr(row, field_name)
        if value not in (None, ""):
            return value
    return None if field_name.endswith("_date") or field_name == "annuity_valid_until" else ""


def choose_status(rows: list[RawMember]) -> tuple[str, list[str]]:
    warnings: list[str] = []
    statuses = {row.status for row in rows if row.status}

    # A ADIMPLENTES é considerada a base principal quando o associado está nela.
    principal = sorted(
        [row for row in rows if row.source_sheet == "ADIMPLENTES"],
        key=lambda r: r.source_priority,
        reverse=True,
    )
    if principal:
        selected = principal[0].status or STATUS_ACTIVE
    elif any(row.source_sheet == "ASSOCIADOS ATIVOS" for row in rows):
        selected = STATUS_ACTIVE
    elif any(row.source_sheet == "FALTA DE PAGAMENTO" for row in rows):
        selected = STATUS_PENDING
    elif any(row.source_sheet == "INADIMPLENTES" for row in rows):
        selected = STATUS_OVERDUE
    elif any(row.source_sheet == "NÃO CONCLUINTE" for row in rows):
        selected = STATUS_INACTIVE
    else:
        selected = STATUS_ACTIVE

    if len(statuses) > 1:
        warnings.append(
            "Situação divergente entre abas: " + ", ".join(sorted(statuses))
            + f"; adotado '{selected}' pela regra de precedência"
        )
    return selected, warnings


def cluster_rows(
    rows: list[RawMember],
    cpf_collisions: set[str],
    membership_collisions: set[str],
) -> list[list[RawMember]]:
    """Agrupa a mesma pessoa sem usar CPF/matrícula reconhecidamente conflitantes."""

    clusters: list[list[RawMember]] = []
    email_index: dict[str, int] = {}
    cpf_index: dict[str, int] = {}
    membership_index: dict[str, int] = {}
    name_birth_index: dict[tuple[str, date], int] = {}
    name_index: dict[str, list[int]] = defaultdict(list)

    def add_indexes(cluster_id: int, row: RawMember) -> None:
        if row.email_valid and row.email:
            email_index.setdefault(row.email, cluster_id)
        if row.cpf_valid and row.cpf and row.cpf not in cpf_collisions:
            cpf_index.setdefault(row.cpf, cluster_id)
        if row.membership_number and row.membership_number not in membership_collisions:
            membership_index.setdefault(row.membership_number, cluster_id)
        if row.normalized_name and row.birth_date:
            name_birth_index.setdefault((row.normalized_name, row.birth_date), cluster_id)
        if row.normalized_name and cluster_id not in name_index[row.normalized_name]:
            name_index[row.normalized_name].append(cluster_id)

    ordered = sorted(rows, key=lambda r: r.source_priority, reverse=True)

    for row in ordered:
        candidates: set[int] = set()
        if row.email_valid and row.email in email_index:
            candidates.add(email_index[row.email])
        if row.cpf_valid and row.cpf not in cpf_collisions and row.cpf in cpf_index:
            candidates.add(cpf_index[row.cpf])
        if (
            row.membership_number
            and row.membership_number not in membership_collisions
            and row.membership_number in membership_index
        ):
            candidates.add(membership_index[row.membership_number])
        if row.normalized_name and row.birth_date:
            key = (row.normalized_name, row.birth_date)
            if key in name_birth_index:
                candidates.add(name_birth_index[key])

        # Nome exato é fallback útil para linhas históricas sem e-mail/CPF.
        if not candidates and row.normalized_name:
            same_name_clusters = name_index.get(row.normalized_name, [])
            if len(same_name_clusters) == 1:
                candidates.add(same_name_clusters[0])

        # Se múltiplos identificadores apontam para clusters diferentes, não
        # unimos automaticamente: cria-se novo cluster e o conflito será revisado.
        if len(candidates) == 1:
            cluster_id = next(iter(candidates))
            clusters[cluster_id].append(row)
            add_indexes(cluster_id, row)
        else:
            cluster_id = len(clusters)
            clusters.append([row])
            add_indexes(cluster_id, row)

    return clusters


def consolidate(rows: list[RawMember]) -> tuple[list[ConsolidatedMember], set[str], set[str]]:
    cpf_collisions = detect_cpf_collisions(rows)
    membership_collisions = detect_membership_collisions(rows)
    clusters = cluster_rows(rows, cpf_collisions, membership_collisions)

    members: list[ConsolidatedMember] = []
    for cluster in clusters:
        member = ConsolidatedMember(rows=cluster)
        member.membership_number = choose_field(cluster, "membership_number") or ""
        member.full_name = choose_field(cluster, "full_name") or ""
        member.birth_date = choose_field(cluster, "birth_date")
        member.cpf = choose_field(cluster, "cpf") or ""
        member.cpf_raw = choose_field(cluster, "cpf_raw") or ""
        member.cpf_valid = bool(member.cpf and cpf_is_valid(member.cpf))
        member.gender = choose_field(cluster, "gender") or ""
        member.city = choose_field(cluster, "city") or ""
        member.state = choose_field(cluster, "state") or ""
        member.phone = choose_field(cluster, "phone") or ""
        member.email = choose_field(cluster, "email") or ""
        member.email_valid = bool(member.email and EMAIL_RE.match(member.email))
        member.education_level = choose_field(cluster, "education_level") or ""
        member.member_category = choose_field(cluster, "member_category") or ""
        member.university = choose_field(cluster, "university") or ""
        member.annuity_valid_until = choose_field(cluster, "annuity_valid_until")
        member.association_status, status_warnings = choose_status(cluster)
        member.warnings.extend(status_warnings)

        cluster_cpfs = {r.cpf for r in cluster if r.cpf_valid and r.cpf}
        cluster_numbers = {r.membership_number for r in cluster if r.membership_number}
        cluster_emails = {r.email for r in cluster if r.email_valid and r.email}
        cluster_names = {r.normalized_name for r in cluster if r.normalized_name}

        if any(cpf in cpf_collisions for cpf in cluster_cpfs):
            member.identity_conflict = True
            member.warnings.append(
                "CPF aparece associado a nomes diferentes na planilha"
            )

        if len(cluster_cpfs) > 1:
            member.identity_conflict = True
            member.warnings.append(
                "Mais de um CPF válido foi encontrado para a mesma pessoa"
            )

        # Matrícula não autentica o associado. Se ela for ambígua, não
        # adivinhamos qual é a correta: deixamos o campo vazio para revisão
        # administrativa e preservamos a identidade segura por e-mail + CPF.
        membership_has_collision = any(
            number in membership_collisions
            for number in cluster_numbers
        )
        if membership_has_collision or len(cluster_numbers) > 1:
            member.membership_number = ""
            if membership_has_collision:
                member.warnings.append(
                    "Matrícula conflitante entre pessoas; não será importada "
                    "automaticamente"
                )
            if len(cluster_numbers) > 1:
                member.warnings.append(
                    "Mais de uma matrícula encontrada para a mesma pessoa; "
                    "nenhuma será importada automaticamente"
                )

        if len(cluster_emails) > 1:
            member.identity_conflict = True
            member.warnings.append(
                "Mais de um e-mail válido encontrado para a mesma pessoa"
            )

        if len(cluster_names) > 1:
            # Pequenas diferenças de grafia e aliases conhecidos são aceitos.
            names = [r.full_name for r in cluster if r.full_name]
            if (
                len(names) >= 2
                and not all(
                    similar_names(names[0], other)
                    for other in names[1:]
                )
            ):
                member.identity_conflict = True
                member.warnings.append(
                    "Nomes divergentes foram agrupados por outro identificador"
                )

        if member.identity_conflict:
            member.migration_state = MIGRATION_CONFLICT
        elif not member.email_valid:
            member.migration_state = MIGRATION_NO_EMAIL
        elif not member.cpf_valid:
            member.migration_state = MIGRATION_REVIEW
            member.warnings.append(
                "CPF inválido ou incompleto; o registro não será migrado "
                "até a correção, pois o primeiro acesso exige e-mail + CPF"
            )
        else:
            member.migration_state = MIGRATION_READY

        members.append(member)

    members.sort(key=lambda m: (m.full_name.upper(), m.email))
    return members, cpf_collisions, membership_collisions


# -----------------------------------------------------------------------------
# Django / banco
# -----------------------------------------------------------------------------
def setup_django(settings_module: str):
    if str(PROJECT_ROOT) not in sys.path:
        sys.path.insert(0, str(PROJECT_ROOT))
    os.environ.setdefault("DJANGO_SETTINGS_MODULE", settings_module)

    try:
        import django
        django.setup()
    except Exception as exc:
        raise RuntimeError(
            "Não foi possível inicializar o Django. Execute este script na raiz "
            "do projeto ABASC e confira DJANGO_SETTINGS_MODULE."
        ) from exc

    from django.conf import settings
    from django.contrib.auth import get_user_model
    from django.db import connection, transaction

    return settings, get_user_model(), connection, transaction


def model_field_names(User) -> set[str]:
    return {
        field.name
        for field in User._meta.get_fields()
        if getattr(field, "concrete", False)
    }


def set_if_field(user, field_names: set[str], name: str, value: Any) -> bool:
    if name not in field_names:
        return False
    setattr(user, name, value)
    return True


def value_is_empty(value: Any) -> bool:
    return value is None or value == ""


def user_indexes(User):
    by_email: dict[str, list[Any]] = defaultdict(list)
    by_cpf: dict[str, list[Any]] = defaultdict(list)
    by_membership: dict[str, list[Any]] = defaultdict(list)

    for user in User.objects.all().iterator():
        email = clean_text(getattr(user, "email", "")).lower()
        if email:
            by_email[email].append(user)
        cpf = only_digits(getattr(user, "cpf", ""))
        if len(cpf) == 11:
            by_cpf[cpf].append(user)
        membership = clean_text(getattr(user, "membership_number", ""))
        if membership:
            by_membership[membership].append(user)
    return by_email, by_cpf, by_membership


def find_existing_user(member: ConsolidatedMember, indexes) -> tuple[Any | None, list[Any]]:
    by_email, by_cpf, by_membership = indexes
    candidates: dict[Any, Any] = {}

    if member.email:
        for user in by_email.get(member.email.lower(), []):
            candidates[user.pk] = user
    cpf_digits = only_digits(member.cpf)
    if len(cpf_digits) == 11:
        for user in by_cpf.get(cpf_digits, []):
            candidates[user.pk] = user
    if member.membership_number:
        for user in by_membership.get(member.membership_number, []):
            candidates[user.pk] = user

    values = list(candidates.values())
    if len(values) == 1:
        return values[0], values
    return None, values


def member_to_user_values(member: ConsolidatedMember, field_names: set[str]) -> dict[str, Any]:
    first_name, last_name = split_name(member.full_name)
    values: dict[str, Any] = {
        "username": member.email.lower(),
        "email": member.email.lower(),
        "first_name": first_name,
        "last_name": last_name,
        "role": "associate",
        "association_status": member.association_status,
        "is_active": member.association_status != STATUS_INACTIVE,
        "is_staff": False,
        "is_superuser": False,
        "phone": member.phone,
        "cpf": only_digits(member.cpf) if member.cpf_valid else "",
        "city": member.city,
        "state": member.state,
        "membership_number": member.membership_number or None,
    }

    optional_values = {
        # Campos planejados para a atualização do modelo. O script só os usa
        # quando eles realmente existirem no User.
        "birth_date": member.birth_date,
        "gender": member.gender,
        "education_level": member.education_level,
        "member_category": member.member_category,
        "university": member.university,
        "annuity_valid_until": member.annuity_valid_until,
        "legacy_imported": True,
        "onboarding_completed": False,
        "migration_status": (
            "ready" if member.migration_state == MIGRATION_READY else "needs_review"
        ),
    }
    values.update(optional_values)
    return {key: value for key, value in values.items() if key in field_names}


def migrate_to_django(
    members: list[ConsolidatedMember],
    *,
    settings_module: str,
    commit: bool,
    update_existing: bool,
    overwrite: bool,
) -> dict[str, int]:
    settings, User, connection, transaction = setup_django(settings_module)
    fields = model_field_names(User)
    indexes = user_indexes(User)

    expected_future_fields = {
        "birth_date",
        "gender",
        "education_level",
        "member_category",
        "university",
        "annuity_valid_until",
        "legacy_imported",
        "onboarding_completed",
        "migration_status",
    }
    missing_future = sorted(expected_future_fields - fields)
    if missing_future:
        print(
            "[AVISO] O User ainda não possui alguns campos planejados: "
            + ", ".join(missing_future)
        )
        print("        A migração continuará somente com os campos que já existem.")

    db_host = str(connection.settings_dict.get("HOST", ""))
    supabase_db = "supabase" in db_host.lower() or "pooler" in db_host.lower()
    auth_configured = bool(
        getattr(settings, "SUPABASE_URL", "")
        and getattr(settings, "SUPABASE_SERVICE_ROLE_KEY", "")
    )
    print(f"[BANCO] Backend: {connection.vendor}")
    print(f"[BANCO] Parece Supabase PostgreSQL: {'sim' if supabase_db else 'não/indeterminado'}")
    print(f"[AUTH] Supabase Auth configurado no backend: {'sim' if auth_configured else 'não'}")
    print("[AUTH] Nenhum usuário Auth será criado ou convidado por este script.")

    stats: dict[str, int] = defaultdict(int)

    # Se não houver --commit, fazemos todo o preflight sem abrir transação de escrita.
    for member in members:
        member.db_action = ""
        member.db_user_id = ""

        # Somente registros aptos ao primeiro acesso podem ser gravados.
        # PRECISA_REVISAO, SEM_EMAIL_VALIDO e DADO_CONFLITANTE ficam apenas
        # no relatório até a base ser corrigida.
        if member.migration_state != MIGRATION_READY:
            member.db_action = MIGRATION_SKIPPED
            stats[MIGRATION_SKIPPED] += 1
            continue

        existing, candidates = find_existing_user(member, indexes)
        if len(candidates) > 1:
            member.migration_state = MIGRATION_CONFLICT
            member.identity_conflict = True
            member.warnings.append(
                "E-mail/CPF/matrícula apontam para mais de um usuário já existente no banco"
            )
            member.db_action = MIGRATION_SKIPPED
            stats[MIGRATION_SKIPPED] += 1
            continue

        if existing is not None:
            member.db_user_id = str(existing.pk)
            if not update_existing:
                member.migration_state = MIGRATION_EXISTING
                member.db_action = MIGRATION_EXISTING
                stats[MIGRATION_EXISTING] += 1
                continue

            values = member_to_user_values(member, fields)
            changed: list[str] = []
            for field_name, incoming in values.items():
                if field_name in {"username", "email"}:
                    # Não alteramos identidade primária de usuário existente.
                    continue
                current = getattr(existing, field_name, None)
                if overwrite or value_is_empty(current):
                    if not value_is_empty(incoming) and current != incoming:
                        setattr(existing, field_name, incoming)
                        changed.append(field_name)

            # Usuário histórico que já existe também deve ser associado comum,
            # mas não rebaixamos presidente/executivo silenciosamente.
            current_role = getattr(existing, "role", "associate")
            if current_role not in {"president", "executive"} and "role" in fields:
                if existing.role != "associate":
                    existing.role = "associate"
                    changed.append("role")

            if commit and changed:
                existing.save(update_fields=sorted(set(changed)))
            member.db_action = MIGRATION_UPDATED if changed else MIGRATION_EXISTING
            stats[member.db_action] += 1
            continue

        if not member.email_valid:
            member.db_action = MIGRATION_SKIPPED
            stats[MIGRATION_SKIPPED] += 1
            continue

        values = member_to_user_values(member, fields)
        user = User(**values)
        user.set_unusable_password()

        if commit:
            try:
                with transaction.atomic():
                    user.save()
            except Exception as exc:
                member.db_action = MIGRATION_ERROR
                member.warnings.append(f"Erro ao salvar: {type(exc).__name__}: {exc}")
                stats[MIGRATION_ERROR] += 1
                continue

            member.db_user_id = str(user.pk)
            # Atualiza índices em memória para a própria execução continuar idempotente.
            indexes[0][user.email.lower()].append(user)
            cpf_digits = only_digits(getattr(user, "cpf", ""))
            if len(cpf_digits) == 11:
                indexes[1][cpf_digits].append(user)
            membership = clean_text(getattr(user, "membership_number", ""))
            if membership:
                indexes[2][membership].append(user)
            member.db_action = MIGRATION_CREATED
        else:
            member.db_action = "CRIARIA"
        stats[member.db_action] += 1

    return stats


# -----------------------------------------------------------------------------
# Relatório
# -----------------------------------------------------------------------------
def date_text(value: date | None) -> str:
    return value.isoformat() if value else ""


def write_report(path: Path, members: list[ConsolidatedMember]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    columns = [
        "resultado",
        "acao_banco",
        "id_usuario",
        "nome",
        "email",
        "cpf",
        "matricula",
        "situacao_associativa",
        "categoria",
        "nascimento",
        "sexo_genero",
        "telefone",
        "municipio",
        "uf",
        "escolaridade",
        "universidade",
        "validade_anuidade",
        "abas_origem",
        "avisos",
    ]

    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns, delimiter=";")
        writer.writeheader()
        for member in members:
            writer.writerow({
                "resultado": member.migration_state,
                "acao_banco": member.db_action,
                "id_usuario": member.db_user_id,
                "nome": member.full_name,
                "email": member.email,
                "cpf": member.cpf or member.cpf_raw,
                "matricula": member.membership_number,
                "situacao_associativa": member.association_status,
                "categoria": member.member_category,
                "nascimento": date_text(member.birth_date),
                "sexo_genero": member.gender,
                "telefone": member.phone,
                "municipio": member.city,
                "uf": member.state,
                "escolaridade": member.education_level,
                "universidade": member.university,
                "validade_anuidade": date_text(member.annuity_valid_until),
                "abas_origem": member.source_sheets,
                "avisos": " | ".join(dict.fromkeys(member.warnings)),
            })


def print_summary(
    raw_rows: list[RawMember],
    members: list[ConsolidatedMember],
    cpf_collisions: set[str],
    membership_collisions: set[str],
    stats: dict[str, int] | None = None,
) -> None:
    by_state: dict[str, int] = defaultdict(int)
    by_assoc_status: dict[str, int] = defaultdict(int)
    by_category: dict[str, int] = defaultdict(int)

    for member in members:
        by_state[member.migration_state] += 1
        by_assoc_status[member.association_status] += 1
        by_category[member.member_category or "não informada"] += 1

    print("\n" + "=" * 72)
    print("RESUMO DA MIGRAÇÃO ABASC")
    print("=" * 72)
    print(f"Linhas úteis lidas nas abas: {len(raw_rows)}")
    print(f"Pessoas consolidadas:        {len(members)}")
    print(f"CPFs conflitantes:           {len(cpf_collisions)}")
    print(f"Matrículas conflitantes:     {len(membership_collisions)}")

    print("\nPossibilidade de migração:")
    for key in (
        MIGRATION_READY,
        MIGRATION_REVIEW,
        MIGRATION_NO_EMAIL,
        MIGRATION_CONFLICT,
        MIGRATION_EXISTING,
    ):
        if by_state.get(key):
            print(f"  {key:<24} {by_state[key]}")

    print("\nSituação associativa resultante:")
    for key in (STATUS_ACTIVE, STATUS_PENDING, STATUS_OVERDUE, STATUS_INACTIVE):
        if by_assoc_status.get(key):
            print(f"  {key:<12} {by_assoc_status[key]}")

    print("\nCategorias:")
    for key, value in sorted(by_category.items(), key=lambda item: (-item[1], item[0])):
        print(f"  {key:<18} {value}")

    if stats:
        print("\nRegra de segurança:")
        print("  somente PRONTO_PARA_MIGRAR pode ser criado/atualizado")
        print("  matrícula ambígua é omitida, nunca escolhida por adivinhação")
        print("\nAções no banco:")
        for key, value in sorted(stats.items()):
            print(f"  {key:<24} {value}")

    if cpf_collisions:
        print("\n[ATENÇÃO] Há CPFs ligados a nomes diferentes. Eles foram bloqueados")
        print("          para evitar vincular uma pessoa à identidade de outra.")

    print("=" * 72)


# -----------------------------------------------------------------------------
# CLI
# -----------------------------------------------------------------------------
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Analisa e migra a base histórica de associados da ABASC."
    )
    parser.add_argument(
        "--planilha",
        type=Path,
        default=DEFAULT_XLSX,
        help="Caminho do XLSX. Padrão: membros/data.xlsx",
    )
    parser.add_argument(
        "--relatorio",
        type=Path,
        default=DEFAULT_REPORT,
        help="CSV de auditoria. Padrão: membros/relatorio_migracao.csv",
    )
    parser.add_argument(
        "--settings",
        default=DEFAULT_SETTINGS_MODULE,
        help="Módulo settings do Django. Padrão: abasc_mvp1.settings",
    )
    parser.add_argument(
        "--commit",
        action="store_true",
        help="Grava no banco. Sem esta opção o script somente simula.",
    )
    parser.add_argument(
        "--update-existing",
        action="store_true",
        help="Preenche campos vazios de usuários já existentes.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Com --update-existing, permite sobrescrever campos preenchidos.",
    )
    parser.add_argument(
        "--somente-planilha",
        action="store_true",
        help="Não inicializa Django; apenas consolida a planilha e gera relatório.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    planilha = args.planilha.expanduser().resolve()
    relatorio = args.relatorio.expanduser().resolve()

    if args.overwrite and not args.update_existing:
        raise SystemExit("--overwrite exige --update-existing.")
    if args.somente_planilha and args.commit:
        raise SystemExit("--somente-planilha não pode ser usado junto com --commit.")

    print(f"[ARQUIVO] {planilha}")
    print(f"[MODO] {'GRAVAÇÃO' if args.commit else 'SIMULAÇÃO / DRY-RUN'}")

    raw_rows = load_members(planilha)
    members, cpf_collisions, membership_collisions = consolidate(raw_rows)

    stats: dict[str, int] | None = None
    if not args.somente_planilha:
        try:
            stats = migrate_to_django(
                members,
                settings_module=args.settings,
                commit=args.commit,
                update_existing=args.update_existing,
                overwrite=args.overwrite,
            )
        except RuntimeError as exc:
            if args.commit:
                raise
            print(f"[AVISO] {exc}")
            print("        O relatório da planilha será gerado sem consultar o banco Django.")

    write_report(relatorio, members)
    print_summary(raw_rows, members, cpf_collisions, membership_collisions, stats)
    print(f"\n[RELATÓRIO] {relatorio}")

    if not args.commit:
        print("\nNenhuma alteração foi gravada. Para migrar de verdade, execute:")
        print("    python membros/migracao.py --commit")
    else:
        print("\nMigração concluída. Nenhuma senha ou conta do Supabase Auth foi criada.")
        print("O próximo passo é habilitar o fluxo de Primeiro acesso / onboarding.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())